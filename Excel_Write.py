import openpyxl
path = "Excel_Write.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
column = sheet.max_column
print(rows)
print(column)
for r in range(1,8):
    for c in range(1,9):
        sheet.cell(row=r,column=c).value="Amit Singh"

workbook.save(path)